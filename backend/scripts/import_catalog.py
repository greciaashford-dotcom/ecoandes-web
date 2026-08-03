"""
Catalog reconciliation importer (Excel-driven).

Sources (in /app/backend/data):
  - precios_profesionales.xlsx  (MASTER: FAMILIA, Código, Descripción, Formato, IVA, Origen, precio pro s/IVA, EAN)
  - precios_web.xlsx            (Código, Descripción, Formato, Grupo, PVP s/IVA, EAN, peso en kg, precio pro s/IVA)

Behaviour:
  - Each Excel row = one SKU = one format/variation.
  - SKUs grouped into base products by stripping trailing digits of the code (AMG1/AMG500/AMG100 -> AMG).
  - Products in DB whose base-prefix is NOT in the Excel master are ARCHIVED (moved to products_archive, removed from products).
  - Matched products keep their id, image_url, gallery, translations, created_at, flags.
  - Prices stored SIN IVA. vat_rate from PRO file (sesame IVA "2" -> 4 provisional).

Usage:
  python -m scripts.import_catalog            # DRY-RUN (no writes), prints report
  python -m scripts.import_catalog --commit   # apply changes
"""
import asyncio
import re
import sys
import uuid
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from core.config import db  # noqa: E402
from core.utils import slugify  # noqa: E402

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
PRO_FILE = DATA_DIR / "precios_profesionales.xlsx"
WEB_FILE = DATA_DIR / "precios_web.xlsx"

SESAME_IVA_FIX = {}  # Pelado sésamo keeps Excel IVA value (2%) per client confirmation


def base_code(sku: str) -> str:
    """Group key: strip trailing digits (AMG500 -> AMG)."""
    return re.sub(r"\d+$", "", str(sku).strip().upper())


def norm_format(fmt) -> str:
    """Normalise '5kg'/'1 Kg' -> '5 kg' / '1 kg'; '250 g' stays."""
    s = str(fmt or "").strip()
    m = re.match(r"^\s*([\d.,]+)\s*(kg|g|gr)\s*$", s, re.IGNORECASE)
    if not m:
        return s
    num = m.group(1).replace(",", ".").rstrip("0").rstrip(".") if "." in m.group(1) else m.group(1)
    unit = m.group(2).lower()
    unit = "kg" if unit == "kg" else "g"
    return f"{num} {unit}"


def f(v, default=0.0):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return default


def load_pro():
    wb = openpyxl.load_workbook(PRO_FILE, data_only=True)
    ws = wb["HOJA DE PEDIDO"]
    out = {}
    for r in range(5, ws.max_row + 1):
        cod = ws.cell(r, 3).value
        if not cod or str(cod).strip() in ("Código", "CÓDIGO"):
            continue
        cod = str(cod).strip()
        iva = ws.cell(r, 6).value
        out[cod] = {
            "family": (ws.cell(r, 1).value or "").strip(),
            "desc": (ws.cell(r, 4).value or "").strip(),
            "fmt": ws.cell(r, 5).value,
            "iva": SESAME_IVA_FIX.get(cod, int(iva) if iva is not None else 10),
            "origin": (ws.cell(r, 7).value or "").strip(),
            "price_pro": f(ws.cell(r, 8).value),
            "ean": str(ws.cell(r, 12).value or "").strip(),
        }
    return out


def load_web():
    wb = openpyxl.load_workbook(WEB_FILE, data_only=True)
    ws = wb["Hoja1"]
    out = {}
    for r in range(2, ws.max_row + 1):
        cod = ws.cell(r, 1).value
        if not cod or str(cod).strip() in ("Código", "CÓDIGO"):
            continue
        cod = str(cod).strip()
        out[cod] = {
            "desc": (ws.cell(r, 2).value or "").strip(),
            "fmt": ws.cell(r, 3).value,
            "grupo": (ws.cell(r, 4).value or "").strip(),
            "price_pro": f(ws.cell(r, 5).value),
            "pvp": f(ws.cell(r, 9).value),
            "ean": str(ws.cell(r, 12).value or "").strip(),
            "weight": f(ws.cell(r, 13).value),
        }
    return out


def derive_name(descriptions):
    """Clean product name: cut each description at the FIRST size token (removes
    format + pack noise), then pick the shortest tidy name (most stripped)."""
    cleaned = set()
    for d in descriptions:
        n = re.split(r"\s+[\d.,]+\s*(?:kg|g|gr)\b", d, maxsplit=1, flags=re.IGNORECASE)[0]
        n = re.sub(r"\([^)]*\)", "", n)            # drop parentheticals
        n = re.sub(r"\s{2,}", " ", n).strip(" -.")
        if n:
            cleaned.add(n)
    if not cleaned:
        return "Producto"
    # shortest = least leftover noise; tie-break alphabetically for determinism
    return sorted(cleaned, key=lambda s: (len(s), s))[0]


def norm_name(s: str) -> str:
    """Normalised product name for fuzzy image carry-over matching."""
    import unicodedata
    s = unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode("ascii").lower()
    s = re.sub(r"\b(bio|sin gluten|sg|eco|ecologico|ecologica)\b", " ", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


async def build():
    pro = load_pro()
    web = load_web()
    all_skus = set(pro) | set(web)

    groups = defaultdict(list)
    for sku in all_skus:
        groups[base_code(sku)].append(sku)

    # existing products -> map (a) base prefix and (b) normalised name -> product (preserve image/gallery/id/...)
    existing = {}
    existing_by_name = {}
    async for p in db.products.find({}, {"_id": 0}):
        prefixes = {base_code(p.get("sku", ""))}
        for v in (p.get("variations") or []):
            prefixes.add(base_code(v.get("sku", "")))
        for pref in prefixes:
            if pref and pref not in existing:
                existing[pref] = p
        nm = norm_name(p.get("name", ""))
        if nm and nm not in existing_by_name:
            existing_by_name[nm] = p

    products = []
    used_prior_ids = set()
    for gcode, skus in groups.items():
        members = []
        descs = []
        family = origin = ""
        vat = 10
        for sku in skus:
            p = pro.get(sku, {})
            w = web.get(sku, {})
            desc = p.get("desc") or w.get("desc") or ""
            descs.append(desc)
            family = family or p.get("family") or ""
            origin = origin or p.get("origin") or ""
            if p.get("iva"):
                vat = p["iva"]
            weight = w.get("weight") or 0.0
            fmt = norm_format(p.get("fmt") or w.get("fmt"))
            price_retail = w.get("pvp", 0.0)
            price_pro = p.get("price_pro") or w.get("price_pro") or 0.0
            members.append({
                "sku": sku,
                "name": fmt,
                "price_retail": round(price_retail, 2),
                "price_professional": round(price_pro, 2),
                "stock": 999,
                "image_url": "",
                "weight_kg": round(weight, 3),
                "is_bulk": weight > 1.0,
                "ean": p.get("ean") or w.get("ean") or "",
                "available_retail": sku in web,
                "available_professional": sku in pro,
            })
        # order variations by weight asc
        members.sort(key=lambda m: m["weight_kg"])
        name = derive_name(descs)
        # base prices = smallest available format
        retail_opts = [m["price_retail"] for m in members if m["available_retail"] and m["price_retail"] > 0]
        pro_opts = [m["price_professional"] for m in members if m["available_professional"] and m["price_professional"] > 0]
        base_retail = min(retail_opts) if retail_opts else (members[0]["price_retail"] if members else 0.0)
        base_pro = min(pro_opts) if pro_opts else (members[0]["price_professional"] if members else 0.0)

        prior = existing.get(gcode) or existing_by_name.get(norm_name(name))
        # avoid reusing the same existing product for two different new groups
        if prior and prior["id"] in used_prior_ids:
            prior = None
        doc = {
            "sku": members[0]["sku"] if members else gcode,
            "name": name,
            "category": family or "General",
            "origin_country": origin,
            "vat_rate": vat,
            "price_retail": base_retail,
            "price_professional": base_pro,
            "variations": members,
            "stock": 999,
            "active": True,
        }
        if prior:
            used_prior_ids.add(prior["id"])
            doc["id"] = prior["id"]
            doc["slug"] = prior.get("slug") or slugify(name)
            # Migración SEO: si el producto fue renombrado al nombre legacy de la
            # web antigua, el Excel NO debe revertirlo (name/slug quedan blindados).
            if prior.get("legacy_name_applied"):
                doc["name"] = prior["name"]
                doc["legacy_name_applied"] = True
            doc["image_url"] = prior.get("image_url", "")
            doc["gallery"] = prior.get("gallery", []) or []
            doc["created_at"] = prior.get("created_at") or datetime.now(timezone.utc).isoformat()
            doc["featured"] = prior.get("featured", False)
            doc["best_seller"] = prior.get("best_seller", False)
            for keep in ("translations", "description", "short_description", "highlights",
                         "description_blocks", "nutrition", "badges", "tech_sheet", "seo",
                         "slug_aliases", "previous_name"):
                if prior.get(keep):
                    doc[keep] = prior[keep]
            doc["_action"] = "update"
        else:
            doc["id"] = str(uuid.uuid4())
            doc["slug"] = slugify(name)
            doc["image_url"] = ""
            doc["gallery"] = []
            doc["created_at"] = datetime.now(timezone.utc).isoformat()
            doc["_action"] = "create"
        doc["_gcode"] = gcode
        products.append(doc)

    # ensure unique slugs across the resulting catalog
    seen_slugs = {}
    for p in products:
        base_slug = p["slug"] or slugify(p["name"])
        s = base_slug
        if s in seen_slugs:
            s = f"{base_slug}-{p['_gcode'].lower()}"
        seen_slugs[s] = True
        p["slug"] = s

    # archive existing products NOT matched to any new group (by id)
    to_archive = []
    async for p in db.products.find({}, {"_id": 0}):
        if p.get("id") not in used_prior_ids:
            to_archive.append(p)

    return products, to_archive


async def main(commit=False):
    products, to_archive = await build()
    creates = [p for p in products if p["_action"] == "create"]
    updates = [p for p in products if p["_action"] == "update"]

    print("=" * 70)
    print(f"RECONCILIACIÓN DE CATÁLOGO  ({'COMMIT' if commit else 'DRY-RUN'})")
    print("=" * 70)
    print(f"Productos base resultantes : {len(products)}")
    print(f"  - A CREAR  : {len(creates)}")
    print(f"  - A SYNC   : {len(updates)}")
    print(f"  - Formatos totales        : {sum(len(p['variations']) for p in products)}")
    print(f"Productos a ARCHIVAR (fuera de Excel): {len(to_archive)}")
    if to_archive:
        print("  ej.:", [p.get("name") for p in to_archive[:10]])
    print("\nEjemplos a crear:", [f"{p['name']} ({len(p['variations'])} fmt, IVA {p['vat_rate']}%)" for p in creates[:6]])
    # sanity: vat distribution
    vat_dist = defaultdict(int)
    for p in products:
        vat_dist[p["vat_rate"]] += 1
    print("Distribución IVA por producto:", dict(vat_dist))

    if not commit:
        print("\n(DRY-RUN: no se escribió nada. Ejecuta con --commit para aplicar.)")
        return

    now = datetime.now(timezone.utc).isoformat()
    # archive
    if to_archive:
        for p in to_archive:
            p["archived_at"] = now
        await db.products_archive.insert_many([{k: v for k, v in p.items()} for p in to_archive])
        ids = [p["id"] for p in to_archive]
        await db.products.delete_many({"id": {"$in": ids}})
    # upsert products
    for p in products:
        p.pop("_action", None)
        p.pop("_gcode", None)
        p["updated_at"] = now
        await db.products.replace_one({"id": p["id"]}, p, upsert=True)
    total = await db.products.count_documents({})
    print(f"\nOK. Catálogo aplicado. Productos en BD ahora: {total}")
    print(f"Archivados en products_archive: {len(to_archive)}")


if __name__ == "__main__":
    asyncio.run(main(commit="--commit" in sys.argv))
