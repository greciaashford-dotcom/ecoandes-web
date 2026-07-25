"""WhatsApp contact leads: public capture + admin management with Excel export."""
import re
import uuid
from datetime import datetime, timezone
from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from core.auth import require_admin
from core.config import db

router = APIRouter(prefix="/api", tags=["whatsapp"])

PHONE_RE = re.compile(r"^[+]?[\d\s\-().]{6,20}$")


class WhatsappLeadCreate(BaseModel):
    name: str = Field(..., min_length=2, max_length=120)
    phone: str = Field(..., min_length=6, max_length=25)


def _clean_phone(phone: str) -> str:
    return re.sub(r"[^\d+]", "", phone.strip())


# ---------- Public: capture lead before opening WhatsApp ----------
@router.post("/whatsapp-leads")
async def create_whatsapp_lead(payload: WhatsappLeadCreate):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=422, detail="El nombre es obligatorio")
    if not PHONE_RE.match(payload.phone.strip()):
        raise HTTPException(status_code=422, detail="Teléfono inválido")

    phone = _clean_phone(payload.phone)
    digits = re.sub(r"\D", "", phone)
    if len(digits) < 6:
        raise HTTPException(status_code=422, detail="Teléfono inválido")

    now = datetime.now(timezone.utc).isoformat()
    existing = await db.whatsapp_leads.find_one({"phone": phone})
    if existing:
        # Same phone contacting again: refresh name and increment counter
        await db.whatsapp_leads.update_one(
            {"phone": phone},
            {"$set": {"name": name, "last_contact_at": now},
             "$inc": {"contact_count": 1}},
        )
        return {"ok": True, "already": True}

    await db.whatsapp_leads.insert_one({
        "id": str(uuid.uuid4()),
        "name": name,
        "phone": phone,
        "contact_count": 1,
        "created_at": now,
        "last_contact_at": now,
    })
    return {"ok": True, "already": False}


# ---------- Admin: list ----------
@router.get("/admin/whatsapp-leads")
async def list_whatsapp_leads(user: dict = Depends(require_admin)):
    leads = await db.whatsapp_leads.find({}, {"_id": 0}).sort("last_contact_at", -1).to_list(5000)
    return {"leads": leads, "total": len(leads)}


# ---------- Admin: delete ----------
@router.delete("/admin/whatsapp-leads/{lead_id}")
async def delete_whatsapp_lead(lead_id: str, user: dict = Depends(require_admin)):
    res = await db.whatsapp_leads.delete_one({"id": lead_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    return {"ok": True}


# ---------- Admin: export to Excel ----------
@router.get("/admin/whatsapp-leads/export")
async def export_whatsapp_leads(user: dict = Depends(require_admin)):
    leads = await db.whatsapp_leads.find({}, {"_id": 0}).sort("last_contact_at", -1).to_list(10000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads WhatsApp"

    headers = ["Nombre", "Teléfono", "Nº contactos", "Primer contacto", "Último contacto"]
    ws.append(headers)
    # Style header
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="4A6B4D", end_color="4A6B4D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for lead in leads:
        ws.append([
            lead.get("name", ""),
            lead.get("phone", ""),
            lead.get("contact_count", 1),
            (lead.get("created_at") or "")[:19].replace("T", " "),
            (lead.get("last_contact_at") or "")[:19].replace("T", " "),
        ])

    # Column widths
    widths = [30, 20, 14, 22, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"leads-whatsapp-ecoandes-{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
